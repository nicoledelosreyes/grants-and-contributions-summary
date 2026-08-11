"""
populate_dv_alldata.py
--------------------------
Populates the "DV - All Data" sheet in "G&C Extraction Updated.xlsx" with:
  - Column A: WBS Prefix (first 8 characters of WBS code)
  - Column B: Program Name (from Program Mapping via WBS code)
  - Column C: TPP Risk Assessment (from Program Mapping col C, matched to col B)
  - Column D: Recipient Name (from Supplier sheet via Journal Entry number)
  - Column E: Dollar value (from SAPUI5 Export column I)

Entries with the same Recipient Name AND the same WBS prefix are combined
into a single row with their dollar values summed. Entries with the same
Recipient Name but a different WBS prefix appear as separate rows.

  - A SUBTOTAL row at the bottom that respects filters

File Structure Requirements:
Ensure that populate_dv_alldata.py is in the same folder as the source file and master workbook. 
The source file must be named "G&C programms and recipients data april 20, 2026.xlsx". 
The master workbook must be named "G&C Extraction.xlsx".

Usage:
Place this script in the same folder as "G&C Extraction Updated.xlsx" and run:
    python populate_dv_alldata.py
"""

import os
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

INPUT_FILE = "G&C Extraction Updated.xlsx"
OUTPUT_FILE = "G&C Extraction Updated.xlsx"
SOURCE_FILE = "G&C programms and recipients data april 20, 2026.xlsx"
DOLLAR_FORMAT = '#,##0.00_0 "CAD";* -#,##0.00_0 "CAD"'
NO_RISK_ASSESSMENT = "[No TPP Risk Assessment]"

SPECIAL_PREFIX = "94-03-02"

# ── Special case recipients ───────────────────────────────────────────────────
# These recipients have no WBS code in the source data and must be hardcoded.
# Add new entries here as { recipient_name: (wbs_prefix, tpp_name) }
SPECIAL_CASE_RECIPIENTS = {
    "BUCTOUCHE MICMAC BAND": (
        "94-03-02",
        "Aboriginal Fisheries Strategy - Negotiation and Implementation of Fisheries Agreements (AFS / NIFA)",
    ),
}


def get_tpp_name(wbs_full: str, prefix_map: dict) -> str:
    wbs_24 = str(wbs_full).strip()[:24]
    if wbs_24[:8] == SPECIAL_PREFIX:
        if wbs_24.endswith("-P-01-00-000-A01"):
            return "Aboriginal Fisheries Strategy - Allocation Transfer Program (AFS / ATP)"
        else:
            return "Aboriginal Fisheries Strategy - Negotiation and Implementation of Fisheries Agreements (AFS / NIFA)"
    if wbs_24 == "":
        return "[No WBS code provided]"
    return prefix_map.get(wbs_24[:8], f"[No mapping for WBS code: {wbs_24[:8]}]")


def build_prefix_map(ws) -> dict:
    # Col A (index 0) = prefix key, Col B (index 1) = TPP Name
    mapping = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        key = str(row[0]).strip() if row[0] is not None else ""
        val = str(row[1]).strip() if row[1] is not None else ""
        if key and key != "None":
            mapping[key] = val
    return mapping


def build_program_to_risk_map(ws) -> dict:
    """{ program_name (col B) → risk_assessment_name (col C) }"""
    mapping = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        program = str(row[1]).strip() if row[1] is not None else ""
        risk    = str(row[2]).strip() if row[2] is not None else ""
        if program and program != "None" and risk and risk != "None":
            mapping[program] = risk
    return mapping


def build_risk_assessment_map(ws) -> dict:
    """{ risk_assessment_name (col A) → risk_level (col B) }
    Reads consecutive rows starting at row 3."""
    mapping = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        key = str(row[0]).strip() if row[0] is not None else ""
        val = str(row[1]).strip() if row[1] is not None else ""
        if key and key != "None":
            mapping[key] = val
    return mapping


def build_supplier_map(ws) -> dict:
    # Col A (index 0) = JE number, Col D (index 3) = Supplier name
    mapping = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        key = str(row[0]).strip() if row[0] is not None else ""
        val = str(row[3]).strip() if row[3] is not None else ""
        if key and key != "None":
            mapping[key] = val
    return mapping


def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    input_path = os.path.join(script_dir, INPUT_FILE)
    output_path = os.path.join(script_dir, OUTPUT_FILE)

    if not os.path.exists(input_path):
        raise FileNotFoundError(f"Could not find '{INPUT_FILE}' in {script_dir}")

    source_path = os.path.join(script_dir, SOURCE_FILE)
    if not os.path.exists(source_path):
        raise FileNotFoundError(f"Could not find '{SOURCE_FILE}' in {script_dir}")

    print(f"Loading: {input_path}")
    wb = load_workbook(input_path)

    print(f"Loading source data: {source_path}")
    wb_src = load_workbook(source_path, read_only=True, data_only=True)

    for name in ["Program Mapping", "DV - All Data", "TPP Risk Assessments"]:
        if name not in wb.sheetnames:
            raise ValueError(f"Sheet '{name}' not found in {INPUT_FILE}. Available: {wb.sheetnames}")

    for name in ["SAPUI5 Export", "Supplier"]:
        if name not in wb_src.sheetnames:
            raise ValueError(f"Sheet '{name}' not found in {SOURCE_FILE}. Available: {wb_src.sheetnames}")

    prefix_map           = build_prefix_map(wb["Program Mapping"])
    program_to_risk_map  = build_program_to_risk_map(wb["Program Mapping"])
    risk_assessment_map  = build_risk_assessment_map(wb["TPP Risk Assessments"])
    supplier_map         = build_supplier_map(wb_src["Supplier"])

    ws_out = wb["DV - All Data"]
    ws_src = wb_src["SAPUI5 Export"]

    # ── Clear all existing data from row 2 downward ───────────────────────────
    for row in ws_out.iter_rows(min_row=2):
        for cell in row:
            cell.value = None
            cell.font = Font()
            cell.fill = PatternFill("none")
            cell.alignment = Alignment()
            cell.number_format = "General"

    # ── Write headers ─────────────────────────────────────────────────────────
    headers = ["WBS", "Program Name", "TPP Risk Assessment", "Recipient Name", "Dollar Values", "Risk Level", "Receiving from multiple programs?", "How many programs in total?"]
    for col, header in enumerate(headers, 1):
        cell = ws_out.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

    # ── Collect and group rows ────────────────────────────────────────────────
    # Key: (wbs_prefix, tpp_name, recipient)  →  summed dollar value
    grouped = defaultdict(float)
    key_order = []
    unmatched_tpp = 0
    unmatched_supplier = 0
    special_case_hits = defaultdict(float)

    for src_row in ws_src.iter_rows(min_row=2, values_only=True):
        # Col C = index 2: Journal Entry number → Recipient Name
        je_number = str(src_row[2]).strip() if src_row[2] is not None else ""
        recipient = supplier_map.get(je_number, f"[Supplier not found: {je_number}]")
        if recipient.startswith("["):
            unmatched_supplier += 1

        # Col I = index 8: Dollar value
        raw_dollar = src_row[8]
        if raw_dollar is None:
            dollar_val = 0.0
        elif isinstance(raw_dollar, (int, float)):
            dollar_val = float(raw_dollar)
        else:
            try:
                dollar_val = float(str(raw_dollar).replace(",", "").strip())
            except (ValueError, AttributeError, TypeError):
                dollar_val = 0.0

        # ── Special case: recipient with no WBS in source data ───────────────
        if recipient in SPECIAL_CASE_RECIPIENTS:
            wbs_prefix, tpp_name = SPECIAL_CASE_RECIPIENTS[recipient]
            special_case_hits[recipient] += dollar_val
        else:
            # Normal path: resolve WBS and TPP from source
            # Col M = index 12: WBS code
            wbs_raw = str(src_row[12]).strip() if src_row[12] is not None else ""
            wbs_prefix = wbs_raw[:8] if wbs_raw else ""
            tpp_name = get_tpp_name(wbs_raw, prefix_map)
            if tpp_name.startswith("["):
                unmatched_tpp += 1

        key = (wbs_prefix, tpp_name, recipient)
        if key not in grouped:
            key_order.append(key)
        grouped[key] += dollar_val

    # ── Build recipient → set of unique programs map ────────────────────────
    recipient_programs = defaultdict(set)
    for (wbs_prefix, tpp_name, recipient) in key_order:
        recipient_programs[recipient].add(tpp_name)

    # ── Sort by dollar value descending ──────────────────────────────────────
    key_order.sort(key=lambda k: grouped[k], reverse=True)

    # ── Write grouped rows ────────────────────────────────────────────────────
    out_row = 2
    for key in key_order:
        wbs_prefix, tpp_name, recipient = key
        dollar_val = grouped[key]
        risk_assessment = program_to_risk_map.get(tpp_name, NO_RISK_ASSESSMENT)
        risk_level = risk_assessment_map.get(risk_assessment, "[N/A]")

        num_programs = len(recipient_programs[recipient])
        multi_program = "Yes" if num_programs > 1 else "No"
        program_count = num_programs if num_programs > 1 else "N/A"

        ws_out.cell(row=out_row, column=1, value=wbs_prefix)
        ws_out.cell(row=out_row, column=2, value=tpp_name)
        ws_out.cell(row=out_row, column=3, value=risk_assessment)
        ws_out.cell(row=out_row, column=4, value=recipient)
        dollar_cell = ws_out.cell(row=out_row, column=5, value=dollar_val)
        dollar_cell.number_format = DOLLAR_FORMAT
        ws_out.cell(row=out_row, column=6, value=risk_level)
        ws_out.cell(row=out_row, column=7, value=multi_program)
        ws_out.cell(row=out_row, column=8, value=program_count)

        out_row += 1

    total_rows = out_row - 2
    data_end = total_rows + 1
    print(f"Source rows processed: {len(grouped)}  →  {total_rows} grouped rows written")
    if special_case_hits:
        print(f"Special case recipients applied:")
        for name, total in special_case_hits.items():
            print(f"  {name}: {total:,.2f}")
    if unmatched_tpp:
        print(f"TPP Name not matched: {unmatched_tpp} rows")
    if unmatched_supplier:
        print(f"Supplier not matched:  {unmatched_supplier} rows")

    # ── SUBTOTAL row — updates automatically when filters are applied ─────────
    subtotal_row = total_rows + 3
    label = ws_out.cell(row=subtotal_row, column=4, value="TOTAL (visible rows)")
    label.font = Font(bold=True)
    label.alignment = Alignment(horizontal="right")

    subtotal = ws_out.cell(row=subtotal_row, column=5, value=f"=SUBTOTAL(9,E2:E{data_end})")
    subtotal.number_format = DOLLAR_FORMAT
    subtotal.font = Font(bold=True)

    fill = PatternFill("solid", start_color="BDD7EE", end_color="BDD7EE")
    label.fill = fill
    subtotal.fill = fill

    ws_out.auto_filter.ref = f"A1:H{data_end}"
    ws_out.column_dimensions["A"].width = 15
    ws_out.column_dimensions["B"].width = 40
    ws_out.column_dimensions["C"].width = 40
    ws_out.column_dimensions["D"].width = 70
    ws_out.column_dimensions["E"].width = 20
    ws_out.column_dimensions["F"].width = 15
    ws_out.column_dimensions["G"].width = 30
    ws_out.column_dimensions["H"].width = 25

    # ── Write total dollar values per TPP Risk Assessment to TPP Risk Assessments col C ──
    ws_tpp = wb["TPP Risk Assessments"]

    # Sum dollar values by risk assessment name from the grouped data
    tpp_totals = defaultdict(float)
    for key, dollar_val in grouped.items():
        _, tpp_name, _ = key
        risk_assessment = program_to_risk_map.get(tpp_name, NO_RISK_ASSESSMENT)
        if risk_assessment != NO_RISK_ASSESSMENT:
            tpp_totals[risk_assessment] += dollar_val

    # Write totals to col C of TPP Risk Assessments, matching on col A name
    tpp_written = 0
    for row in ws_tpp.iter_rows(min_row=3, values_only=False):
        name_cell = row[0]
        if name_cell.value is None:
            continue
        name = str(name_cell.value).strip()
        dollar_cell = row[2]  # col C
        if name in tpp_totals:
            dollar_cell.value = tpp_totals[name]
            dollar_cell.number_format = DOLLAR_FORMAT
            tpp_written += 1
        else:
            dollar_cell.value = None

    print(f"TPP Risk Assessment totals written: {tpp_written}")

    wb.save(output_path)
    print(f"\nSaved: {output_path}")


if __name__ == "__main__":
    main()
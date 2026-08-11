"""
populate_tpp_risk_assessments.py
------------------------------
Reads F26:F37 and F39 from the Summary sheet of each source file,
general comments and additional comments from each RF1-RF12 sheet,
and C46 from Summary for the overall RF comment (column Q).

Column layout (each program occupies 2 rows):

  Col A     Program name
  Col B     Overall Risk Level (from K39 of Summary sheet)
  Col C     Dollar Value (reserved blank)
  Col D     Onboarded to ECGS (Yes/No)
  Col E-P   RF1-RF12: score on score row, comment on comment row
  Col Q     Overall: score on score row, overall comment on comment row

File Structure Requirements:

Ensure that populate_tpp_risk_assessments.py is in the same folder as the source files and master workbook. 
All source files must be in the "Valid TTP Risk Assessments" subfolder. 
The master workbook must be named "G&C Extraction.xlsx".

Usage:
    python populate_tpp_risk_assessments.py
"""

import os
import glob
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side


# ── Paths ──────────────────────────────────────────────────────────────────────
BASE_DIR    = r"C:\Users\DELOSREYESN\OneDrive - DFO-MPO\Documents\G&C Recipient Audits\populate_tpp_risk_assessments"
SOURCE_DIR  = os.path.join(BASE_DIR, "Valid TTP Risk Assessments")
MASTER_FILE = os.path.join(BASE_DIR, "G&C Extraction.xlsx")

# ── Fills ──────────────────────────────────────────────────────────────────────
FILL_GREEN  = PatternFill("solid", fgColor="92D050")
FILL_YELLOW = PatternFill("solid", fgColor="FFCC00")
FILL_RED    = PatternFill("solid", fgColor="FF1D1D")
FILL_GREY   = PatternFill("solid", fgColor="AEAAAA")
FILL_NONE   = PatternFill("none")

# ── Alignments ─────────────────────────────────────────────────────────────────
CENTER   = Alignment(horizontal="center", vertical="center")
LEFT_TOP = Alignment(horizontal="left",   vertical="top", wrap_text=True)

# ── Borders ────────────────────────────────────────────────────────────────────
THIN = Side(style="thin")
NO   = Side(style=None)

BORDER_ALL         = Border(left=THIN, top=THIN, right=THIN, bottom=THIN)
BORDER_NONE        = Border(left=NO,   top=NO,   right=NO,   bottom=NO)
BORDER_MERGE_TOP   = Border(left=THIN, top=THIN, right=THIN, bottom=NO)
BORDER_MERGE_BOT   = Border(left=THIN, top=NO,   right=THIN, bottom=THIN)

# ── ECGS onboarded acronyms ────────────────────────────────────────────────────
ECGS_ACRONYMS = {
    "OMP", "PSF", "SEP", "YSSC", "Sustainable Fisheries", "T.Buck", "QFF", "AFF", "BCSRIF", "CEBP",
    "FHSCP", "FRCP", "MEQ", "NCAG", "OCCS", "OFSCP", "SFSF", "WPRICP",
}

# ── RF mappings ────────────────────────────────────────────────────────────────
RF_GENERAL_CELLS = {
    "RF1":  "C66",
    "RF2":  "C59",
    "RF3":  "C29",
    "RF4":  "C42",
    "RF5":  "C20",
    "RF6":  "C37",
    "RF7":  "C26",
    "RF8":  "C22",
    "RF9":  "C25",
    "RF10": "C65",
    "RF11": "C35",
    "RF12": "C45",
}

RF_ADDITIONAL_RANGES = {
    "RF1":  ("F69", "F73"),
    "RF2":  ("F62", "F66"),
    "RF3":  ("F32", "F36"),
    "RF4":  ("F45", "F49"),
    "RF5":  ("F23", "F27"),
    "RF6":  ("F40", "F44"),
    "RF7":  ("F29", "F33"),
    "RF8":  ("F25", "F29"),
    "RF9":  ("F28", "F32"),
    "RF10": ("F68", "F72"),
    "RF11": ("F38", "F42"),
    "RF12": ("F48", "F52"),
}

ADDITIONAL_HEADERS = [
    "Mitigation strategies",
    "Reporting and monitoring",
    "Responsibility for reporting and monitoring",
    "Timing",
    "Resource requirements",
]


# ── Helpers ────────────────────────────────────────────────────────────────────
def get_risk_level(summary_ws):
    raw = summary_ws["K39"].value
    if raw is None:
        return "", FILL_NONE
    val = str(raw).strip()
    val_lower = val.lower()
    if val_lower in ("high", "élevé"):
        return "High", FILL_RED
    elif val_lower in ("medium", "moyen"):
        return "Medium", FILL_YELLOW
    elif val_lower in ("low", "faible"):
        return "Low", FILL_GREEN
    else:
        return val, FILL_NONE


def get_ecgs_status(title):
    if " - " in title:
        acronym = title.rsplit(" - ", 1)[-1].strip()
    else:
        acronym = title.strip()
    return "Yes" if acronym in ECGS_ACRONYMS else "No"


def format_value(raw):
    """Return (display_value, comment_fill) for a score."""
    if raw is None or raw == "":
        return 0, FILL_NONE
    if isinstance(raw, str) and raw.strip().upper() == "N/A":
        return "N/A", FILL_GREY
    try:
        num = float(raw)
    except (ValueError, TypeError):
        return str(raw), FILL_NONE

    display = round(num, 1)
    if 1 <= num < 1.5:
        return display, FILL_GREEN
    elif 1.5 <= num < 2.5:
        return display, FILL_YELLOW
    elif num >= 2.5:
        return display, FILL_RED
    else:
        return display, FILL_NONE


# Writing prompts to treat as blank
PLACEHOLDER_TEXTS = {
    "(e.g., hiring additional staff within 6 months using generic work descriptions and statements of merit criteria, Gs&Cs training has been arranged for program officers, succession plans have been established)",
    "(how the mitigation strategies will be monitored and reported on)",
    "(program officer, manager, director, name)",
    "(how frequently will the risk mitigation strategies be tracked and reported to management)",
    "(how much money and time will it take to put mitigation strategies in place and to track the strategies)",
}


def get_comment(val):
    if val is None:
        return ""
    s = str(val).strip()
    if not s or s in PLACEHOLDER_TEXTS:
        return ""
    return s


def build_rf_comment(ws, rf_name):
    """Build comment text with General Comments and Additional Comments headers."""
    general = get_comment(ws[RF_GENERAL_CELLS[rf_name]].value)

    start_cell, end_cell = RF_ADDITIONAL_RANGES[rf_name]
    start_row = int(''.join(filter(str.isdigit, start_cell)))
    end_row   = int(''.join(filter(str.isdigit, end_cell)))
    col       = 6  # column F

    additional_lines = []
    for i, row_num in enumerate(range(start_row, end_row + 1)):
        val = get_comment(ws.cell(row=row_num, column=col).value)
        header = ADDITIONAL_HEADERS[i]
        additional_lines.append(f"{header}: {val if val else '-'}")

    all_blank = all(
        not get_comment(ws.cell(row=r, column=col).value)
        for r in range(start_row, end_row + 1)
    )

    parts = []

    if general:
        parts.append(f"General Comments:\n{general}")
    else:
        parts.append("General Comments: none")

    if not all_blank:
        parts.append("Additional Comments:\n" + "\n".join(additional_lines))
    else:
        parts.append("Additional Comments: none")

    return "\n\n".join(parts)


def build_overall_comment(overall_comment_text):
    """Build overall comment cell text."""
    if overall_comment_text and overall_comment_text != "-":
        return f"General Comments:\n{overall_comment_text}"
    else:
        return "General Comments: none"


def clear_cell(cell):
    """Reset a cell to completely blank."""
    cell.value = None
    cell.fill = FILL_NONE
    cell.border = BORDER_NONE
    cell.alignment = Alignment()


def read_source(filepath):
    wb = load_workbook(filepath, data_only=True)

    if "Summary" in wb.sheetnames:
        summary = wb["Summary"]
    elif "Résumé" in wb.sheetnames:
        summary = wb["Résumé"]
    else:
        raise ValueError(f'No "Summary" or "Résumé" sheet in {os.path.basename(filepath)}')

    title = os.path.splitext(os.path.basename(filepath))[0]
    risk_value, risk_fill = get_risk_level(summary)

    # F26:F37 — 12 RF scores, F39 — overall score
    raw_scores = []
    for row in summary.iter_rows(min_row=26, max_row=37, min_col=6, max_col=6, values_only=True):
        raw_scores.append(row[0])
    raw_scores.append(summary["F39"].value)

    # (display, comment_fill) for each score
    scores = [format_value(r) for r in raw_scores]

    # Build RF comments
    comments = []
    for i in range(1, 13):
        sheet_name = f"RF{i}"
        if sheet_name in wb.sheetnames:
            comments.append(build_rf_comment(wb[sheet_name], sheet_name))
        else:
            comments.append("General Comments: none\n\nAdditional Comments: none")

    # Overall comment
    overall_val = get_comment(summary["C46"].value)
    overall_text = overall_val if overall_val else "-"
    overall_comment = build_overall_comment(overall_text)

    wb.close()
    return title, risk_value, risk_fill, scores, comments, overall_comment


# ── Main ───────────────────────────────────────────────────────────────────────
def main():
    pattern = os.path.join(SOURCE_DIR, "2025-26 TPP Risk Assessment - *.xlsx")
    source_files = sorted(glob.glob(pattern))

    if not source_files:
        print(f"No source files found in:\n  {SOURCE_DIR}")
        return

    print(f"Found {len(source_files)} source files. Reading...")

    entries = []
    for filepath in source_files:
        try:
            title, risk_value, risk_fill, scores, comments, overall_comment = read_source(filepath)
            entries.append((title, risk_value, risk_fill, scores, comments, overall_comment))
            print(f"  ✓ {title}")
        except Exception as e:
            print(f"  ✗ {os.path.basename(filepath)}: {e}")

    print(f"\nWriting {len(entries)} entries to master workbook...")

    master_wb = load_workbook(MASTER_FILE)
    if "TPP Risk Assessments" not in master_wb.sheetnames:
        raise ValueError('"TPP Risk Assessments" sheet not found in master workbook.')

    sheet = master_wb["TPP Risk Assessments"]

    TOTAL_COLS = 17

    # ── Unmerge any existing merged cells in A3:Q downward ─────────────────────
    max_clear_row = 3 + (len(entries) * 2) + 50
    merged_to_remove = [
        str(r) for r in sheet.merged_cells.ranges
        if r.min_row >= 3 and r.min_col >= 1 and r.max_col <= TOTAL_COLS
    ]
    for r in merged_to_remove:
        sheet.unmerge_cells(r)

    # ── Clear A3:Q downward ────────────────────────────────────────────────────
    for r in range(3, max_clear_row + 1):
        for c in range(1, TOTAL_COLS + 1):
            clear_cell(sheet.cell(row=r, column=c))

    # ── Write entries — 2 rows per program ────────────────────────────────────
    for row_idx, (title, risk_value, risk_fill, scores, comments, overall_comment) in enumerate(entries):
        score_row   = 3 + (row_idx * 2)
        comment_row = score_row + 1

        sheet.row_dimensions[score_row].height = 14.5
        sheet.row_dimensions[comment_row].height = 70

        ecgs_value = get_ecgs_status(title)

        # ── Merge columns A–D across both rows ────────────────────────────────
        for col in range(1, 5):
            sheet.merge_cells(
                start_row=score_row, start_column=col,
                end_row=comment_row, end_column=col
            )

        # Col A — program name: centered with wrap, merged borders
        cell = sheet.cell(row=score_row, column=1, value=title)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = FILL_NONE
        cell.border = BORDER_MERGE_TOP
        sheet.cell(row=comment_row, column=1).border = BORDER_MERGE_BOT

        # Col B — Overall Risk Level: centered, coloured, merged borders
        cell = sheet.cell(row=score_row, column=2, value=risk_value)
        cell.alignment = CENTER
        cell.fill = risk_fill
        cell.border = BORDER_MERGE_TOP
        bottom = sheet.cell(row=comment_row, column=2)
        bottom.fill = risk_fill
        bottom.border = BORDER_MERGE_BOT

        # Col C — Dollar Value: blank, centered, merged borders
        cell = sheet.cell(row=score_row, column=3, value=None)
        cell.alignment = CENTER
        cell.fill = FILL_NONE
        cell.border = BORDER_MERGE_TOP
        bottom = sheet.cell(row=comment_row, column=3)
        bottom.fill = FILL_NONE
        bottom.border = BORDER_MERGE_BOT

        # Col D — ECGS Yes/No: centered, no fill, merged borders
        cell = sheet.cell(row=score_row, column=4, value=ecgs_value)
        cell.alignment = CENTER
        cell.fill = FILL_NONE
        cell.border = BORDER_MERGE_TOP
        bottom = sheet.cell(row=comment_row, column=4)
        bottom.fill = FILL_NONE
        bottom.border = BORDER_MERGE_BOT

        # ── Score row: Cols E–P (RF1–RF12 scores), Col Q (overall score) ──────
        for col_idx, (display, comment_fill) in enumerate(scores[:12]):
            cell = sheet.cell(row=score_row, column=col_idx + 5, value=display)
            cell.alignment = CENTER
            cell.fill = comment_fill
            cell.border = BORDER_ALL

        overall_display, overall_fill = scores[12]
        cell = sheet.cell(row=score_row, column=17, value=overall_display)
        cell.alignment = CENTER
        cell.fill = overall_fill
        cell.border = BORDER_ALL

        # ── Comment row: Cols E–P (RF comments), Col Q (overall comment) ──────
        for col_idx, comment in enumerate(comments):
            cell = sheet.cell(row=comment_row, column=col_idx + 5, value=comment)
            cell.alignment = LEFT_TOP
            cell.fill = FILL_NONE
            cell.border = BORDER_ALL

        cell = sheet.cell(row=comment_row, column=17, value=overall_comment)
        cell.alignment = LEFT_TOP
        cell.fill = FILL_NONE
        cell.border = BORDER_ALL

    master_wb.save(MASTER_FILE)
    master_wb.close()

    print(f"\nDone! Results saved to:\n  {MASTER_FILE}")


if __name__ == "__main__":
    main()
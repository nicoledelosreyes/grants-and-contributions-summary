"""
populate_dv_overview.py
-----------------------
Sets up the "DV - Overview" sheet in "G&C Extraction Updated.xlsx".

  - B3: dropdown of TPP Risk Assessment names
  - C3: accumulated total for all recipients under selected TPP
  - D3: risk level text + conditional color
  - B6+: Recipient Name  }  via FILTER() formula, spills automatically
  - C6+: Individual Dollar Value  }  when B3 selection changes

Run once. After that, changing B3 updates everything automatically.

File Structure Requirements:
Ensure that populate_dv_overview.py is in the same folder as the master workbook. 
The master workbook must be named "G&C Extraction.xlsx".

Place in the same folder as "G&C Extraction.xlsx" and run:
    python populate_dv_overview.py
"""

import os
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import FormulaRule

INPUT_FILE = "G&C Extraction Updated.xlsx"
PROMPT = "-- Select a TPP --"

# ── Set to True to rebuild the B3 dropdown, False to skip ────────────────────
REBUILD_DROPDOWN = False

RISK_COLORS = {
    "high":   "FF6565",
    "élevé":  "FF6565",
    "medium": "FFCC00",
    "moyen":  "FFCC00",
    "low":    "92D050",
    "faible": "92D050",
}

DOLLAR_FORMAT = '#,##0.00_0 "CAD";* -#,##0.00_0 "CAD"'


def build_tpp_to_program(ws_pm):
    """{ tpp_name (col C) → program_name (col B) }"""
    mapping = {}
    for row in ws_pm.iter_rows(min_row=2, values_only=True):
        program  = str(row[1]).strip() if row[1] is not None else ""
        tpp_name = str(row[2]).strip() if row[2] is not None else ""
        if tpp_name and tpp_name != "None":
            mapping[tpp_name] = program
    return mapping


def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    input_path = os.path.join(script_dir, INPUT_FILE)

    if not os.path.exists(input_path):
        raise FileNotFoundError(f"Could not find '{INPUT_FILE}' in {script_dir}")

    wb = load_workbook(input_path)

    for name in ["TPP Risk Assessments", "Program Mapping", "DV - All Data"]:
        if name not in wb.sheetnames:
            raise ValueError(f"Sheet '{name}' not found. Available: {wb.sheetnames}")

    if "DV - Overview" not in wb.sheetnames:
        ws_overview = wb.create_sheet("DV - Overview")
    else:
        ws_overview = wb["DV - Overview"]

    ws_tpp = wb["TPP Risk Assessments"]
    ws_pm  = wb["Program Mapping"]
    ws_all = wb["DV - All Data"]

    # ── Collect TPP names and risk levels ─────────────────────────────────────
    tpp_names   = []
    risk_levels = []
    row_num = 3
    while True:
        name_cell = ws_tpp.cell(row=row_num, column=1)
        risk_cell = ws_tpp.cell(row=row_num, column=2)
        if name_cell.value is None:
            break
        tpp_names.append(str(name_cell.value).strip())
        risk_levels.append(str(risk_cell.value).strip() if risk_cell.value is not None else "")
        row_num += 2

    # ── Count rows in source sheets for formula ranges ────────────────────────
    data_last_row = sum(1 for row in ws_all.iter_rows(min_row=2, values_only=True) if row[0] is not None) + 1
    pm_last_row   = sum(1 for row in ws_pm.iter_rows(min_row=2, values_only=True) if row[0] is not None) + 1

    # ── Helper columns Z (26) and Y (25) ─────────────────────────────────────
    # Z: prompt + TPP names  → dropdown source
    # Y: "" + risk levels    → D3 lookup source
    ws_overview.cell(row=1, column=26, value=PROMPT)
    ws_overview.cell(row=1, column=25, value="")
    for i, (name, risk) in enumerate(zip(tpp_names, risk_levels)):
        ws_overview.cell(row=i + 2, column=26, value=name)
        ws_overview.cell(row=i + 2, column=25, value=risk)

    total_helper_rows = len(tpp_names) + 1

    # ── B3: dropdown ──────────────────────────────────────────────────────────
    if REBUILD_DROPDOWN:
        if ws_overview["B3"].value is None:
            ws_overview["B3"] = PROMPT
        dv = DataValidation(
            type="list",
            formula1=f"'DV - Overview'!$Z$1:$Z${total_helper_rows}",
            allow_blank=True,
            showDropDown=False,
        )
        dv.sqref = "B3"
        ws_overview.add_data_validation(dv)
        print("Dropdown rebuilt in B3.")
    else:
        print("Dropdown skipped (REBUILD_DROPDOWN = False).")

    # ── Program name lookup formula (B3 TPP → Program Mapping col C→B) ───────
    program_lookup = (
        f"IFERROR(INDEX('Program Mapping'!$B$2:$B${pm_last_row},"
        f"MATCH(B3,'Program Mapping'!$C$2:$C${pm_last_row},0)),\"\")"
    )

    # ── C3: accumulated total ─────────────────────────────────────────────────
    ws_overview["C3"] = (
        f'=IFERROR(SUM(FILTER(\'DV - All Data\'!$D$2:$D${data_last_row},'
        f'\'DV - All Data\'!$B$2:$B${data_last_row}={program_lookup},0)),0)'
    )
    ws_overview["C3"].number_format = DOLLAR_FORMAT
    ws_overview["C3"].font = Font(bold=True)
    ws_overview["C3"].alignment = Alignment(horizontal="right", vertical="center")

    # ── D3: risk level text + conditional color ───────────────────────────────
    ws_overview["D3"] = (
        f'=IF(OR(B3="{PROMPT}",B3=""),"-",'
        f'IFERROR(INDEX(\'DV - Overview\'!$Y$2:$Y${total_helper_rows},'
        f'MATCH(B3,\'DV - Overview\'!$Z$2:$Z${total_helper_rows},0)),"-"))'
    )

    ws_overview.conditional_formatting.add("D3", FormulaRule(
        formula=['OR(D3="-",D3="")'], fill=PatternFill("none")
    ))
    for label, hex_color in RISK_COLORS.items():
        ws_overview.conditional_formatting.add("D3", FormulaRule(
            formula=[f'LOWER(D3)="{label}"'],
            fill=PatternFill("solid", start_color=hex_color, end_color=hex_color),
        ))

    # ── B6: FILTER for Recipient Name (col C of DV - All Data) ───────────────
    ws_overview["B6"] = (
        f'=IFERROR(FILTER(\'DV - All Data\'!$C$2:$C${data_last_row},'
        f'\'DV - All Data\'!$B$2:$B${data_last_row}={program_lookup},'
        f'"-- Select a TPP above --"),"")'
    )

    # ── C6: FILTER for Dollar Value (col D of DV - All Data) ─────────────────
    ws_overview["C6"] = (
        f'=IFERROR(FILTER(\'DV - All Data\'!$D$2:$D${data_last_row},'
        f'\'DV - All Data\'!$B$2:$B${data_last_row}={program_lookup},0),0)'
    )
    ws_overview["C6"].number_format = DOLLAR_FORMAT

    wb.save(input_path)
    print(f"Done. Formulas written to C3, D3, B6, C6.")
    print(f"  {len(tpp_names)} TPP names loaded.")
    print(f"  DV - All Data range: rows 2–{data_last_row}")
    print(f"  Program Mapping range: rows 2–{pm_last_row}")


if __name__ == "__main__":
    main()